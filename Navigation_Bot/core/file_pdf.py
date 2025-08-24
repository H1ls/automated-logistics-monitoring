# scan_ka_from_docs.py
# -*- coding: utf-8 -*-
from __future__ import annotations
from pathlib import Path
from typing import Optional, List
import re
import sys

# 1) Папка с документами (.pdf, .xlsm)
FOLDER = r"C:\Users\Hils\PycharmProjects\pet.project\Navigation_Bot\pdf"


# ---- утилиты ---------------------------------------------------------------

def clean(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\x00", " ").replace("\u00A0", " ").replace("\u2009", " ").replace("\u202F", " ")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def cut_tail_company(s: str) -> str:
    """Обрезаем хвост реквизитов, чтобы КА было только названием."""
    if not s:
        return ""
    s = s.strip()

    # 1) Обрезаем всё после ИНН/КПП/ОГРН/ОКПО
    s = re.split(r"\b(ИНН|КПП|ОГРН|ОКПО|БИК|р/с|к/с|адрес регистрации)\b", s, maxsplit=1, flags=re.I)[0]

    # 2) Если после названия идёт длинная цифра (>=8 подряд), тоже обрезаем
    s = re.split(r"\s\d{8,}", s, maxsplit=1)[0]

    return s.strip(" ,;—-")


# ---- чтение файлов ---------------------------------------------------------

def read_pdf_text(path: Path) -> str:
    """Простой парс текста PDF (PyPDF2). Если либа не установлена — подсказываем."""
    try:
        from PyPDF2 import PdfReader  # pip install PyPDF2
    except Exception as e:
        raise RuntimeError("Установи библиотеку: pip install PyPDF2") from e

    try:
        reader = PdfReader(str(path))
        pages = []
        for p in reader.pages:
            pages.append(p.extract_text() or "")
        return "\n".join(pages)
    except Exception:
        return ""


def read_xlsm_text(path: Path) -> str:
    """Склеиваем значения всех ячеек (openpyxl)."""
    try:
        from openpyxl import load_workbook  # pip install openpyxl
    except Exception as e:
        raise RuntimeError("Установи библиотеку: pip install openpyxl") from e

    try:
        wb = load_workbook(str(path), data_only=True, read_only=True, keep_vba=True)
        parts: List[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                line = "  ".join(str(c).strip() for c in row if c is not None)
                if line:
                    parts.append(line)
        return "\n".join(parts)
    except Exception:
        return ""


def read_any(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        return read_pdf_text(path)
    if path.suffix.lower() == ".xlsm":
        return read_xlsm_text(path)
    return ""


# ---- извлечение КА ---------------------------------------------------------

# 1-я волна: явные метки

RX_KA_1 = re.compile(
    r"(?:Наименование|Наим\.?)\s*[:\-–]?\s*(?P<name>[^\n\r]+?)\s*(?=\b(ИНН|КПП|ОГРН|Перевозчик|Грузоотправитель|Грузополучатель|Наименование)\b|$)",
    re.IGNORECASE)
# 2-я волна: строка с организационно-правовой формой (ООО/АО/ИП и т.п.)
RX_KA_2 = re.compile(r"(?:Контрагент|КА|Заказчик)\s*[:\-–]?\s*(?P<name>[^\n\r]+)", re.IGNORECASE)
RX_KA_3 = re.compile(r"\b(ООО|ПАО|АО|ОАО|ЗАО|ИП)\b[^\n\r]+", re.IGNORECASE)
# Берём ТОЛЬКО первое юрлицо (ООО/АО/ИП…) из строки
# было:
# ORG_ONLY_RX = re.compile(r'\b(?:ООО|ПАО|АО|ОАО|ЗАО|ИП)\b\s*(?:["«][^"»]+["»]|[^,\n\r;]+)', re.IGNORECASE)

# стало:
ORG_ONLY_RX = re.compile(
    r'\b(?:МП\s+)?(?:ООО|ПАО|АО|ОАО|ЗАО|ИП|Общество\s+с\s+ограниченной\s+ответственностью)\b'
    r'\s*(?:["«][^"»]+["»]|[^,\n\r;]+)',
    re.IGNORECASE
)

# ORG_ONLY_RX = re.compile(r'\b(?:ООО|ПАО|АО|ОАО|ЗАО|ИП)\b\s*(?:["«][^"»]+["»]|[^,\n\r;]+)', re.IGNORECASE)
STOP_LINE_RX = re.compile(r'\b(банк|bic|бик|р/с|к/с|тел\.?|факс|e-?mail|почта)\b', re.IGNORECASE)
STOP_TAIL_RX = re.compile(
    r'\b(ИНН|КПП|ОГРН|ОКПО|БИК|р/с|к/с|Перевозчик|Грузоотправитель|Грузополучатель|Наименование)\b',
    re.IGNORECASE
)
SECTION_LABELS = {
    "payer": ["плательщик", "заказчик", "плательщик (заказчик)"],
    "shipper": ["грузоотправитель"],
    "consignee": ["грузополучатель"],
}
STOP_LABELS = [
    "перевозчик", "грузоотправитель", "грузополучатель",
    "плательщик", "банк", "реквизит", "инн/кпп", "юридический адрес",
    "контакт", "тел", "факс"
]


def normalize_org_caption(org: str) -> str:
    """
    Если встретили 'МП Общество с ограниченной ответственностью "Имя"',
    возвращаем 'ООО "Имя"'.
    """
    if not org:
        return org
    m = re.search(
        r'(?:МП\s+)?Общество\s+с\s+ограниченной\s+ответственностью\s*(["«])([^"»]+)(["»])',
        org, flags=re.IGNORECASE
    )
    if m:
        name = m.group(2).strip()
        return f'ООО "{name}"'
    return org


def find_org_in_section(text: str, section_labels: list[str]) -> str | None:
    """
    Находит первую строку с ООО/АО/ИП в пределах 5 строк ПОСЛЕ заголовка секции.
    Пропускает «перевозчик», банковские строки и контакты.
    """
    lines = text.splitlines()
    low = [ln.lower() for ln in lines]

    # где начинается раздел
    start_idx = None
    for i, ln in enumerate(low):
        if any(lbl in ln for lbl in section_labels):
            # игнорируем если это раздел «перевозчик»
            if "перевозчик" in ln:
                continue
            start_idx = i
            break
    if start_idx is None:
        return None

    # в пределах окна ищем строку с ООО/АО/ИП; допускаем «Наименование …»
    for j in range(0, 7):  # заголовок + до 6 следующих строк
        k = start_idx + j
        if k >= len(lines):
            break
        raw = lines[k].strip()
        if not raw:
            continue
        if STOP_LINE_RX.search(raw):
            continue
        # если «Наименование: ...», обрежем хвост и возьмём только первое юрлицо
        m_named = re.search(r'(?:Наименование|Наим\.?)\s*[:\-–]?\s*(.+)', raw, re.IGNORECASE)
        if m_named:
            cand = m_named.group(1)
            m_org = ORG_ONLY_RX.search(cand)
            if m_org:
                return m_org.group(0).strip()
        # иначе пробуем напрямую строку
        m_org = ORG_ONLY_RX.search(raw)
        if m_org:
            return m_org.group(0).strip()
    return None


def first_org_name(s: str) -> str:
    if not s:
        return ""
    m = ORG_ONLY_RX.search(s)
    return m.group(0).strip() if m else s.strip()


def cut_by_stops(s: str) -> str:
    if not s:
        return ""
    return STOP_TAIL_RX.split(s, maxsplit=1)[0].strip()


def extract_ka(text: str) -> str | None:
    t = clean(text)

    # 0) Секция «Плательщик/Заказчик»
    org = find_org_in_section(t, SECTION_LABELS["payer"])
    if org:
        org = normalize_org_caption(org)
        return cut_tail_company(org)

    # 1) Секция «Грузоотправитель»
    org = find_org_in_section(t, SECTION_LABELS["shipper"])
    if org:
        org = normalize_org_caption(org)
        return cut_tail_company(org)

    # 2) Секция «Грузополучатель»
    org = find_org_in_section(t, SECTION_LABELS["consignee"])
    if org:
        org = normalize_org_caption(org)
        return cut_tail_company(org)

    # 3) Наименование … (на той же строке) — с отрезанием по стоп‑словам и только ПЕРВОЕ юрлицо
    m = RX_KA_1.search(t)
    if m:
        cand = cut_by_stops(clean(m.group("name")))
        cand = first_org_name(cand)
        if cand and not STOP_LINE_RX.search(cand):
            return cut_tail_company(cand)

    # 4) Заказчик/Контрагент/КА — на той же строке
    m = RX_KA_2.search(t)
    if m:
        cand = cut_by_stops(clean(m.group("name")))
        cand = first_org_name(cand)
        if cand and not STOP_LINE_RX.search(cand):
            return cut_tail_company(cand)

    # 5) Фолбэк: первая строка с ООО/АО/ИП, но пропускаем «перевозчик» и банковские/контактные строки
    for line in t.splitlines():
        low = line.lower()
        if "перевозчик" in low or STOP_LINE_RX.search(line):
            continue
        m2 = RX_KA_3.search(line)
        if m2:
            cand = cut_by_stops(clean(line))
            cand = first_org_name(cand)
            if cand:
                return cut_tail_company(cand)
    return None


# ---- main ------------------------------------------------------------------

def main():
    folder = Path(FOLDER)
    if not folder.exists():
        print(f"Папка не найдена: {folder}")
        sys.exit(1)

    files = sorted([p for p in folder.iterdir() if p.suffix.lower() in {".pdf", ".xlsm"}])
    if not files:
        print("В папке нет .pdf/.xlsm файлов.")
        sys.exit(0)

    print(f"Нашёл файлов: {len(files)}\n")

    for i, p in enumerate(files, 1):
        text = read_any(p)
        ka = extract_ka(text)
        ka_show = ka if ka else "— не найдено —"
        print(f"{i:02d}. {p.name}\n    КА: {ka_show}\n")


if __name__ == "__main__":
    main()
