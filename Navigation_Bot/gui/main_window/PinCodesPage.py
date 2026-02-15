from __future__ import annotations

import json
import os
import time
from dataclasses import dataclass
from typing import List, Dict, Optional

from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtWidgets import (
    QWidget, QGridLayout, QLabel, QLineEdit, QListWidget, QListWidgetItem,
    QFrame, QPushButton, QHBoxLayout, QVBoxLayout
)

from openpyxl import load_workbook


@dataclass
class PinRow:
    card: str  # столбец A "Номер" (номер карты)
    pin: str  # столбец B "Пин код"
    ts: str  # столбец C "Закреплена за" (номер ТС)
    supplier: str  # столбец D "Кем выдана" (поставщик)

    @property
    def n_card(self) -> str:
        return PinCodesPage.norm(self.card)

    @property
    def n_ts(self) -> str:
        return PinCodesPage.norm(self.ts)


class PinCodesPage(QWidget):
    """
    Локальная страница "Пин коды":
    - читает xlsx (Лист 1)
    - строит json-индекс (для быстрой загрузки) + хранит rows
    - автодополнение:
        * по Номер ТС — подстрока (например "49" -> "Н 498..." и "Р 549...")
        * по номеру карты — подстрока (например "3490" -> "3001034900"...)
    """

    def __init__(self, xlsx_path: str, json_path: str, log_func=None, parent: Optional[QWidget] = None, ):
        super().__init__(parent)
        self.xlsx_path = xlsx_path
        self.json_path = json_path
        self.log = log_func or (lambda *a, **k: None)

        self.rows: List[PinRow] = []
        self._mtime: float = 0.0
        self._idx_by_ts = {}

        self._build_ui()

        # грузим из json (если есть), иначе — из xlsx
        data = self._load_cache()
        if data:
            try:
                self.rows = [PinRow(**r) for r in data.get("rows", [])]
                self._mtime = float(data.get("meta", {}).get("xlsx_mtime", 0.0) or 0.0)
            except Exception:
                self.rows = []

        # if not self.rows:
        #     self._rebuild_from_xlsx(force=True)

        # автопоиск
        self.in_ts.textChanged.connect(self._on_ts_changed)
        self.in_card.textChanged.connect(self._on_card_changed)

        # всё тяжёлое и таймер — после того, как виджет уже создан в Qt
        QTimer.singleShot(0, self._post_init)

        # авто-обновление xlsx -> rows/json
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._refresh_if_needed)
        self._timer.start(60_000)  # раз в минуту

    # ---------------- UI ----------------

    def _fmt_zeros_from_cell(self, cell, default_width: int | None = None) -> str:
        """
        Вернёт строку значения ячейки, сохранив ведущие нули по number_format вида '0000'.
        Если формат не задан, можно применить default_width (например 4 для PIN).
        """
        v = cell.value
        if v is None:
            return ""

        if isinstance(v, str):
            return v.strip()

        try:
            s = str(int(v))
        except Exception:
            return str(v).strip()

        fmt = (cell.number_format or "").strip()
        if fmt and all(ch == "0" for ch in fmt):
            return s.zfill(len(fmt))

        if default_width:
            return s.zfill(default_width)

        return s

    def _reindex(self):
        self._idx_by_ts = {}
        for i, r in enumerate(self.rows):
            nts = r.n_ts
            if not nts:
                continue
            self._idx_by_ts.setdefault(nts, []).append(i)

    def _post_init(self):
        # грузим кэш
        data = self._load_cache()
        if data:
            try:
                self.rows = [PinRow(**r) for r in data.get("rows", [])]
                self._mtime = float(data.get("meta", {}).get("xlsx_mtime", 0.0) or 0.0)
                self._reindex()

            except Exception:
                self.rows = []

        # если кэша нет — строим из xlsx
        if not self.rows:
            self._rebuild_from_xlsx(force=True)

        # запускаем авто-обновление
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._refresh_if_needed)
        self._timer.start(60_000)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(6)

        # 1) верх: Номер ТС и Карта + поля ввода
        grid.addWidget(QLabel("Номер ТС"), 0, 0)
        grid.addWidget(QLabel("Карта"), 0, 1)

        self.in_ts = QLineEdit()
        self.in_ts.setPlaceholderText("Вводи часть номера ТС (пример: 49)")

        self.in_card = QLineEdit()
        self.in_card.setPlaceholderText("Вводи часть номера карты (пример: 3490)")

        grid.addWidget(self.in_ts, 1, 0)
        grid.addWidget(self.in_card, 1, 1)

        root.addLayout(grid)

        # список подсказок (две колонки)
        sugg = QGridLayout()
        sugg.setHorizontalSpacing(12)
        sugg.setVerticalSpacing(6)

        self.ts_list = QListWidget()
        self.ts_list.setMaximumHeight(140)
        self.ts_list.itemClicked.connect(self._on_ts_suggestion_clicked)

        self.card_list = QListWidget()
        self.card_list.setMaximumHeight(140)
        self.card_list.itemClicked.connect(self._on_card_suggestion_clicked)

        sugg.addWidget(QLabel("Подсказки по ТС"), 0, 0)
        sugg.addWidget(QLabel("Подсказки по картам"), 0, 1)
        sugg.addWidget(self.ts_list, 1, 0)
        sugg.addWidget(self.card_list, 1, 1)

        root.addLayout(sugg)

        # разделитель
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        root.addWidget(line)

        # 2) ниже: вывод карта/пин/поставщик
        out = QGridLayout()
        out.setHorizontalSpacing(12)
        out.setVerticalSpacing(6)

        out.addWidget(QLabel("Карта:"), 0, 0)
        self.out_card = QLabel("—")
        self.out_card.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        out.addWidget(self.out_card, 0, 1)

        out.addWidget(QLabel("ПИН:"), 1, 0)
        self.out_pin = QLabel("—")
        self.out_pin.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        out.addWidget(self.out_pin, 1, 1)

        out.addWidget(QLabel("Поставщик:"), 2, 0)
        self.out_supplier = QLabel("—")
        self.out_supplier.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        out.addWidget(self.out_supplier, 2, 1)

        root.addLayout(out)

        # кнопка ручного обновления
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.btn_refresh = QPushButton("Обновить из XLSX")
        self.btn_refresh.clicked.connect(lambda *_: self._rebuild_from_xlsx(force=True))
        btn_row.addWidget(self.btn_refresh)
        root.addLayout(btn_row)

    # ---------------- Helpers ----------------


    @staticmethod
    def norm(s: str) -> str:
        return (s or "").strip().upper().replace(" ", "")

    def _set_result(self, row: Optional[PinRow]):
        if not row:
            self.out_card.setText("—")
            self.out_pin.setText("—")
            self.out_supplier.setText("—")
            return
        self.out_card.setText(row.card or "—")
        self.out_pin.setText(row.pin or "—")
        self.out_supplier.setText(row.supplier or "—")

    def _card_display(self, r: PinRow) -> str:
        sup = (r.supplier or "").strip()
        if sup:
            return f"{r.card} | {sup}"
        return r.card

    # ---------------- Autocomplete ----------------
    def _fill_cards_for_exact_ts(self, nts: str, limit: int = 50):
        self.card_list.clear()
        if not nts:
            return

        idxs = self._idx_by_ts.get(nts, [])
        for idx in idxs[:limit]:
            r = self.rows[idx]
            it = QListWidgetItem(self._card_display(r))

            it.setData(Qt.ItemDataRole.UserRole, idx)  # у тебя уже int-индексы
            self.card_list.addItem(it)

    def _on_ts_changed(self):
        q = self.norm(self.in_ts.text())
        self._fill_ts_suggestions(q)

        #  если есть точное совпадение ТС — покажем все карты по нему
        if q in self._idx_by_ts:
            self._fill_cards_for_exact_ts(q)
        else:
            # иначе оставляем card_list как подсказки по введённой карте (или очищаем)
            # self.card_list.clear()
            pass

        row = self._find_best_by_ts(q)
        if row:
            self._set_result(row)

    def _on_card_changed(self):
        q = self.norm(self.in_card.text())
        self._fill_card_suggestions(q)

        row = self._find_best_by_card(q)
        if row:
            self._set_result(row)

    def _fill_ts_suggestions(self, q: str):
        self.ts_list.clear()
        if not q:
            return

        hits = self._find_rows_by_ts_substring(q, limit=5)
        for idx in hits:
            r = self.rows[idx]
            it = QListWidgetItem(r.ts)
            it.setData(Qt.ItemDataRole.UserRole, idx)
            self.ts_list.addItem(it)

    def _fill_card_suggestions(self, q: str):
        self.card_list.clear()
        if not q:
            return

        hits = self._find_rows_by_card_substring(q, limit=5)
        for idx in hits:
            r = self.rows[idx]
            # it = QListWidgetItem(r.card)
            it = QListWidgetItem(self._card_display(r))

            it.setData(Qt.ItemDataRole.UserRole, idx)  #  храним только int
            self.card_list.addItem(it)

    def _on_ts_suggestion_clicked(self, item: QListWidgetItem):
        idx = item.data(Qt.ItemDataRole.UserRole)
        if idx is None:
            return
        try:
            idx = int(idx)
            r = self.rows[idx]
        except Exception:
            return

        # заполняем поле ТС
        self.in_ts.blockSignals(True)
        self.in_ts.setText(r.ts)
        self.in_ts.blockSignals(False)

        #  обновить список карт для этого ТС
        self._fill_cards_for_exact_ts(r.n_ts)

        # (по желанию) очистить поле карты, чтобы не мешало
        # self.in_card.clear()

        # результат можно либо не показывать, либо показать первый попавшийся
        self._set_result(r)

    def _on_card_suggestion_clicked(self, item: QListWidgetItem):
        idx = item.data(Qt.ItemDataRole.UserRole)
        if idx is None:
            return
        try:
            idx = int(idx)
            r = self.rows[idx]
        except Exception:
            return

        self.in_card.blockSignals(True)
        self.in_card.setText(r.card)
        self.in_card.blockSignals(False)

        self._set_result(r)

    def _find_rows_by_ts_substring(self, q: str, limit: int = 5) -> List[int]:
        res: List[int] = []
        for i, r in enumerate(self.rows):
            if q in r.n_ts:
                res.append(i)
                if len(res) >= limit:
                    break
        return res

    def _find_rows_by_card_substring(self, q: str, limit: int = 5) -> List[int]:
        res: List[int] = []
        for i, r in enumerate(self.rows):
            if q in r.n_card:
                res.append(i)
                if len(res) >= limit:
                    break
        return res

    def _find_best_by_ts(self, q: str) -> Optional[PinRow]:
        if not q:
            return None
        # 1) точное совпадение
        for r in self.rows:
            if r.n_ts == q:
                return r
        # 2) если это просто часть — берём первый hit
        for r in self.rows:
            if q in r.n_ts:
                return r
        return None

    def _find_best_by_card(self, q: str) -> Optional[PinRow]:
        if not q:
            return None
        for r in self.rows:
            if r.n_card == q:
                return r
        for r in self.rows:
            if q in r.n_card:
                return r
        return None

    # ---------------- XLSX / JSON cache ----------------

    def _refresh_if_needed(self):
        try:
            if not os.path.exists(self.xlsx_path):
                return
            mtime = os.path.getmtime(self.xlsx_path)
            if mtime > self._mtime:
                self._rebuild_from_xlsx(force=True)
        except Exception as e:
            self.log(f"❌ PinCodesPage авто-обновление: {e}")

    def _load_cache(self) -> Optional[dict]:
        if not os.path.exists(self.json_path):
            return None
        try:
            with open(self.json_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _save_cache(self):
        try:
            os.makedirs(os.path.dirname(self.json_path), exist_ok=True)
        except Exception:
            pass
        data = {
            "meta": {
                "updated_at": time.time(),
                "xlsx_path": self.xlsx_path,
                "xlsx_mtime": self._mtime,
            },
            "rows": [r.__dict__ for r in self.rows],
        }
        try:
            with open(self.json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.log(f"❌ Не удалось сохранить json-кэш пин-кодов: {e}")

    def _rebuild_from_xlsx(self, force: bool = False):
        if not os.path.exists(self.xlsx_path):
            self.log(f"⚠️ Не найден XLSX: {self.xlsx_path}")
            return

        mtime = os.path.getmtime(self.xlsx_path)
        if (not force) and (mtime <= self._mtime):
            return

        try:
            wb = load_workbook(self.xlsx_path, data_only=True)
            # у тебя "Лист 1" один, но на всякий:
            ws = wb["Лист 1"] if "Лист 1" in wb.sheetnames else wb.active

            # строгая схема колонок:
            # A: "Номер" (карта)
            # B: "Пин код"
            # C: "Закреплена за" (ТС)
            # D: "Кем выдана" (поставщик)

            # проверка заголовков (не обязателен, но полезно)
            hA = str(ws["A1"].value or "").strip()
            hB = str(ws["B1"].value or "").strip()
            hC = str(ws["C1"].value or "").strip()
            hD = str(ws["D1"].value or "").strip()

            if hA.lower() != "номер" or hB.lower() != "пин код":
                # не валим работу — просто лог
                self.log(f"ℹ️ Заголовки A1/B1 не совпали строго: '{hA}' / '{hB}'")

            rows: List[PinRow] = []
            # читаем со 2 строки
            for r in range(2, ws.max_row + 1):
                cell_card = ws.cell(row=r, column=1)  # A
                cell_pin = ws.cell(row=r, column=2)  # B
                cell_ts = ws.cell(row=r, column=3)  # C
                cell_sup = ws.cell(row=r, column=4)  # D

                card_s = self._fmt_zeros_from_cell(cell_card)  # без default_width
                pin_s = self._fmt_zeros_from_cell(cell_pin, default_width=4)  # PIN -> 4 знака

                ts = cell_ts.value
                sup = cell_sup.value

                ts_s = (str(ts).strip() if ts is not None else "")
                sup_s = (str(sup).strip() if sup is not None else "")

                if not card_s and not ts_s:
                    continue

                rows.append(PinRow(card=card_s, pin=pin_s, ts=ts_s, supplier=sup_s))

            self.rows = rows
            self._reindex()

            self._mtime = mtime
            self._save_cache()
            self.log(f"✅ Пин-коды: загружено {len(self.rows)} строк из XLSX")

            # обновим подсказки по текущему вводу
            self._fill_ts_suggestions(self.norm(self.in_ts.text()))
            self._fill_card_suggestions(self.norm(self.in_card.text()))

        except Exception as e:
            self.log(f"❌ Ошибка чтения XLSX пин-кодов: {e}")
