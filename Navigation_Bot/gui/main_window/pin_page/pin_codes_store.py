from __future__ import annotations

import json
import os
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from Navigation_Bot.core.domain.entities.pin_code import PinRow
from Navigation_Bot.core.logging import noop_log, normalize_log_func


@dataclass
class PinCodesData:
    rows: list[PinRow]
    xlsx_mtime: float = 0.0
    pins_loaded: bool = False


class PinCodesStore:
    def __init__(self, xlsx_path: str | Path, json_path: str | Path, log_func: Callable[..., None] | None = None):
        self.xlsx_path = str(xlsx_path)
        self.json_path = str(json_path)
        self.log = normalize_log_func(log_func or noop_log)

    @staticmethod
    def format_cell_value(cell, default_width: int | None = None) -> str:
        value = cell.value
        if value is None:
            return ""

        if isinstance(value, str):
            return value.strip()

        try:
            text = str(int(value))
        except Exception:
            return str(value).strip()

        number_format = (cell.number_format or "").strip()
        if number_format and all(ch == "0" for ch in number_format):
            return text.zfill(len(number_format))

        if default_width:
            return text.zfill(default_width)

        return text

    def load_cache(self) -> PinCodesData | None:
        data = self._load_cache_data()
        if not data:
            return None

        try:
            rows = [PinRow(**row) for row in data.get("rows", [])]
            xlsx_mtime = float(data.get("meta", {}).get("xlsx_mtime", 0.0) or 0.0)
            return PinCodesData(rows=rows,
                                xlsx_mtime=xlsx_mtime,
                                pins_loaded=any(bool(row.pin) for row in rows))
        except Exception:
            return None

    def rebuild_from_xlsx(self, *, current_mtime: float = 0.0, force: bool = False) -> PinCodesData | None:
        if not os.path.exists(self.xlsx_path):
            self.log(f"⚠️ Не найден XLSX: {self.xlsx_path}")
            return None

        xlsx_mtime = os.path.getmtime(self.xlsx_path)
        if (not force) and (xlsx_mtime <= current_mtime):
            return None

        try:
            workbook = load_workbook(self.xlsx_path, data_only=True)
            worksheet = workbook["Лист 1"] if "Лист 1" in workbook.sheetnames else workbook.active

            header_card = str(worksheet["A1"].value or "").strip()
            header_pin = str(worksheet["B1"].value or "").strip()
            if header_card.lower() != "номер" or header_pin.lower() != "пин код":
                self.log(f"ℹ️ Заголовки A1/B1 не совпали строго: '{header_card}' / '{header_pin}'")

            rows: list[PinRow] = []
            for row_idx in range(2, worksheet.max_row + 1):
                card_value = self.format_cell_value(worksheet.cell(row=row_idx, column=1))
                pin_value = self.format_cell_value(worksheet.cell(row=row_idx, column=2), default_width=4)

                ts_raw = worksheet.cell(row=row_idx, column=3).value
                supplier_raw = worksheet.cell(row=row_idx, column=4).value
                ts_value = str(ts_raw).strip() if ts_raw is not None else ""
                supplier_value = str(supplier_raw).strip() if supplier_raw is not None else ""

                if not card_value and not ts_value:
                    continue

                rows.append(PinRow(card=card_value,
                                   pin=pin_value,
                                   ts=ts_value,
                                   supplier=supplier_value))

            result = PinCodesData(rows=rows, xlsx_mtime=xlsx_mtime, pins_loaded=True)
            self.save_cache(result)
            return result

        except Exception as exc:
            self.log(f"❌ Ошибка чтения XLSX пин-кодов: {exc}")
            return None

    def save_cache(self, data: PinCodesData) -> None:
        payload = {"meta": {"updated_at": time.time(),
                            "xlsx_path": str(self.xlsx_path),
                            "xlsx_mtime": data.xlsx_mtime,
                            "contains_pin": False},
                   "rows": [self._cache_row(row) for row in data.rows]}
        try:
            self._write_cache_data(payload)
        except Exception as exc:
            self.log(f"❌ Не удалось сохранить json-кэш пин-кодов: {exc}")

    def _load_cache_data(self) -> dict | None:
        if not os.path.exists(self.json_path):
            return None

        try:
            with open(self.json_path, "r", encoding="utf-8") as file:
                data = json.load(file)
            if self._strip_pins_from_cache_data(data):
                self._write_cache_data(data)
            return data
        except Exception:
            return None

    @staticmethod
    def _strip_pins_from_cache_data(data: dict | None) -> bool:
        if not isinstance(data, dict):
            return False

        changed = False
        rows = data.get("rows", [])
        if not isinstance(rows, list):
            return False

        for row in rows:
            if isinstance(row, dict) and "pin" in row:
                row.pop("pin", None)
                changed = True

        return changed

    @staticmethod
    def _cache_row(row: PinRow) -> dict:
        return {"card": row.card,
                "ts": row.ts,
                "supplier": row.supplier}

    def _write_cache_data(self, data: dict) -> None:
        try:
            os.makedirs(os.path.dirname(self.json_path), exist_ok=True)
        except Exception:
            pass

        with open(self.json_path, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)
