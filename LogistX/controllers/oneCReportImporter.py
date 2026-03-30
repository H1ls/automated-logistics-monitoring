# pet.project\LogistX\controllers\oneCReportImporter.py
from __future__ import annotations

import json
import time
from pathlib import Path

import pyautogui
import pydirectinput as pdi
import pygetwindow as gw
import pyperclip
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class OneCReportImporter:
    def __init__(self, log_func=print):
        self.log = log_func

        # куда сохраняем результат JSON
        self.out_json_path = Path("LogistX/config") / "logistx_sample.json"

        # куда сохраняем “сырой” отчет xlsx
        self.out_xlsx_path = Path("LogistX/config") / "1c_unclosed_races.xlsx"

        # подсказка заголовка окна RDP (поменяешь под себя)
        self.rdp_title_hint = "176.57.78.6:2025 — Подключение к удаленному рабочему столу"
        # запасные варианты (часто встречаются)
        self.rdp_title_fallbacks = ["Подключение к удаленному рабочему столу",
                                    "Remote Desktop Connection", ]

        self._rdp_win = None
        # pyautogui safety
        pyautogui.FAILSAFE = True

    def run(self) -> int:
        """Возвращает кол-во сохранённых строк."""
        if not self._activate_rdp():
            self.log("⚠️ Окно RDP не найдено. Открой RDP и попробуй ещё раз.")
            return 0

        text = self._copy_table_text()
        if not self._looks_like_report(text):
            self.log("❌ Не похоже на отчёт (или таблица не выделилась/не скопировалась).")
            return 0

        header, rows_1c = self._parse_tsv(text)
        if not rows_1c:
            self.log("❌ Не удалось распарсить строки отчёта.")
            return 0

        #  шаг 4: сохраняем xlsx (сырой отчет)
        self._save_xlsx(header, rows_1c)

        # шаг 5: конвертим в нужную структуру и сохраняем json
        rows = self._to_logistx_rows(rows_1c)
        # self._save_json(rows)
        synced_rows = self._sync_json(rows)

        return len(synced_rows)

    # ---- helpers for json sync ----
    @staticmethod
    def _race_key(row: dict) -> str:
        return str(row.get("Рейс", "") or "").strip()

    def _load_existing_json(self) -> list[dict]:
        if not self.out_json_path.exists():
            return []

        try:
            data = json.loads(self.out_json_path.read_text(encoding="utf-8") or "[]")
            if isinstance(data, list):
                return [x for x in data if isinstance(x, dict)]
            if isinstance(data, dict):
                return [data]
            return []
        except Exception as e:
            self.log(f"⚠️ Не удалось прочитать существующий JSON: {e}")
            return []

    def _sync_json(self, imported_rows: list[dict]) -> list[dict]:
        """
        Синхронизация по полю 'Рейс':
        - существующие и снова импортированные рейсы оставляем как есть
        - новые рейсы добавляем
        - исчезнувшие из нового импорта удаляем
        Порядок берём из нового импорта 1С.
        """
        self.out_json_path.parent.mkdir(parents=True, exist_ok=True)

        existing_rows = self._load_existing_json()
        existing_by_race = {}

        for row in existing_rows:
            race = self._race_key(row)
            if race:
                existing_by_race[race] = row

        synced_rows = []
        added = 0
        kept = 0

        seen = set()

        for imported in imported_rows:
            race = self._race_key(imported)
            if not race:
                continue

            if race in seen:
                continue
            seen.add(race)

            old_row = existing_by_race.get(race)
            if old_row is not None:
                synced_rows.append(old_row)
                kept += 1
            else:
                synced_rows.append(imported)
                added += 1

        removed = 0
        imported_races = {self._race_key(r) for r in imported_rows if self._race_key(r)}
        for race in existing_by_race:
            if race not in imported_races:
                removed += 1

        self.out_json_path.write_text(
            json.dumps(synced_rows, ensure_ascii=False, indent=2),
            encoding="utf-8", )

        self.log(f"✅ JSON sync сохранён: {self.out_json_path} | "
                 f"kept={kept}, added={added}, removed={removed}, total={len(synced_rows)}")

        return synced_rows

    # ---- RDP focus + copy ----
    def _find_rdp_window(self):
        # 1) по основному hint
        wins = gw.getWindowsWithTitle(self.rdp_title_hint)
        if wins:
            return wins[0]

        # 2) по fallback’ам
        for hint in self.rdp_title_fallbacks:
            wins = gw.getWindowsWithTitle(hint)
            if wins:
                return wins[0]

        # 3) по частичному совпадению среди всех (на всякий)
        try:
            all_titles = gw.getAllTitles()
            for t in all_titles:
                if not t:
                    continue
                if self.rdp_title_hint in t:
                    w = gw.getWindowsWithTitle(t)
                    if w:
                        return w[0]
        except Exception:
            pass

        return None

    def _activate_rdp(self) -> bool:
        w = self._find_rdp_window()
        if not w:
            return False

        try:
            if getattr(w, "isMinimized", False):
                w.restore()
                time.sleep(0.2)

            # активируем
            w.activate()
            time.sleep(0.25)

            # разворачиваем
            try:
                w.maximize()
                time.sleep(0.25)
            except Exception:
                pass

            self._rdp_win = w
            return True
        except Exception:
            return False

    def _copy_table_text(self) -> str:
        # очищаем буфер
        try:
            pyperclip.copy("")
        except Exception:
            pass

        # координаты центра RDP-окна
        w = self._rdp_win
        if w:
            x = 900
            y = 200
            # x = int(w.left + w.width * 0.5)
            # y = int(w.top + w.height * 0.5)
        else:
            sw, sh = pyautogui.size()
            x, y = sw // 2, sh // 2

        # фокус: лучше двойной клик
        pyautogui.click(x, y)
        time.sleep(0.12)
        pyautogui.click(x, y)
        time.sleep(0.15)

        pdi.PAUSE = 0  # отключаем авто-задержки
        pdi.FAILSAFE = False

        # Выделить всё
        pdi.keyDown("ctrl")
        time.sleep(0.05)
        pdi.press("a")
        time.sleep(0.05)
        pdi.keyUp("ctrl")

        time.sleep(0.15)
        # Копировать
        pdi.keyDown("ctrl")
        time.sleep(0.05)
        pdi.press("c")
        time.sleep(0.05)
        pdi.keyUp("ctrl")

        time.sleep(0.2)

        txt = self._wait_clipboard()
        if txt:
            return txt

        # копировать (вариант 2: часто помогает в RDP/1C)
        pyautogui.hotkey("ctrl", "insert", interval=0.05)
        txt = self._wait_clipboard()
        if txt:
            return txt

        return (pyperclip.paste() or "").strip()

    def _wait_clipboard(self, tries: int = 12, delay: float = 0.12) -> str:
        for _ in range(tries):
            time.sleep(delay)
            try:
                txt = (pyperclip.paste() or "").strip()
            except Exception:
                txt = ""
            if txt:
                return txt
        return ""

    # ---- validation ----
    def _looks_like_report(self, text: str) -> bool:
        if not text or "\t" not in text:
            return False
        t = text.lower()
        return ("тс" in t) and ("рейс" in t)

    # ---- parse ----
    def _parse_tsv(self, text: str) -> tuple[list[str], list[dict]]:
        lines = [ln for ln in text.splitlines() if ln.strip()]
        if not lines:
            return [], []

        header_idx = 0
        for i, ln in enumerate(lines[:12]):
            if "\t" in ln and ("тс" in ln.lower()) and ("рейс" in ln.lower()):
                header_idx = i
                break

        header = [h.strip() for h in lines[header_idx].split("\t")]
        data_lines = lines[header_idx + 1:]

        rows = []
        for ln in data_lines:
            if "\t" not in ln:
                continue
            parts = ln.split("\t")
            if len(parts) < len(header):
                parts += [""] * (len(header) - len(parts))
            obj = {header[i]: parts[i].strip() for i in range(len(header))}
            rows.append(obj)

        return header, rows

    # ---- save xlsx ----
    def _save_xlsx(self, header: list[str], rows: list[dict]) -> None:
        self.out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "1C_Report"

        ws.append(header)
        for r in rows:
            ws.append([r.get(h, "") for h in header])

        # небольшая авто-ширина
        for col_idx, h in enumerate(header, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max([len(str(h or ""))] + [len(str(rr.get(h, "") or "")) for rr in rows[:200]])
            ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))

        wb.save(self.out_xlsx_path)
        self.log(f"✅ XLSX сохранён: {self.out_xlsx_path}")

    # ---- convert ----
    def _to_logistx_rows(self, rows_1c: list[dict]) -> list[dict]:
        def pick(obj, *keys):
            for k in keys:
                if k in obj and str(obj[k]).strip():
                    return obj[k]
            return ""

        out = []
        for r in rows_1c:
            row = {"ТС": pick(r, "ТС"),
                   "Рейс": pick(r, "Рейс"),
                   "Рейс.Пункт отправления": pick(r, "Рейс.Пункт отправления", "Пункт отправления", "Отправление"),
                   "Рейс.Пункт назначения": pick(r, "Рейс.Пункт назначения", "Пункт назначения", "Назначение"),
                   "Плановая дата освобождения разгрузка": pick(r, "Плановая дата освобождения разгрузка",
                                                                "Плановая дата"),
                   "Контрагент": pick(r, "Контрагент"), }

            # фильтр: если ВСЕ поля пустые — пропускаем (убирает хвостовые пустые строки)
            if all(not str(v).strip() for v in row.values()):
                continue

            out.append(row)

        return out

    def _save_json(self, rows: list[dict]) -> None:
        self.out_json_path.parent.mkdir(parents=True, exist_ok=True)
        self.out_json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
        self.log(f"✅ JSON сохранён: {self.out_json_path}")
